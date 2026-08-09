#!/usr/bin/env python3
"""Build the facturación historical store from legacy Excel into local SQLite (D1 schema).

Reads the Provincias yearly invoice ledgers and the Máscaras sales/expenses workbook,
cleans them, and loads a single `transactions` fact table + `expenses` table.

Usage:  python build_history.py [--downloads DIR] [--out history.db]
Output: history.db (SQLite, same schema as D1) + a summary report to stdout.
Re-runnable: the DB is rebuilt from scratch each run.
"""
import argparse, glob, json, os, re, sqlite3, sys
from datetime import datetime, date

HERE = os.path.dirname(os.path.abspath(__file__))
NOW = datetime.utcnow().isoformat(timespec="seconds") + "Z"

# ---------- cleaning helpers ----------

def num(v):
    """Coerce a cell to float, or None. Rejects #REF!, text, blanks."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(".", "").replace(",", ".") if v.count(",") == 1 and v.count(".") > 1 else v.strip()
        # only accept a clean numeric string
        try:
            return float(v.strip())
        except ValueError:
            return None
    return None

YEAR_MIN, YEAR_MAX = 2019, 2027   # plausible operating window; outside → flagged

def as_date(v):
    """Return ISO 'YYYY-MM-DD' if v is a real date, else None.
    Repairs the '19/07/20243' 5-digit-year typo → '19/07/2024'."""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, str):
        s = v.strip()
        m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4,5})$", s)
        if m:
            d_, mo, yr = m.groups()
            if len(yr) == 5:          # stray trailing digit: '20243' -> '2024'
                yr = yr[:4]
            try:
                return datetime(int(yr), int(mo), int(d_)).date().isoformat()
            except ValueError:
                return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                pass
    return None

def year_of(iso):
    return int(iso[:4]) if iso else None

def date_flag(iso):
    """Mark dates outside the plausible operating window for review."""
    y = year_of(iso)
    return "date_suspect" if (y is None or y < YEAR_MIN or y > YEAR_MAX) else None

CHANNEL_PREFIX = re.compile(r"^\s*(ML|IG|FB|TN|WA|WSP|WhatsApp)\s*[:\-]", re.IGNORECASE)

def split_channel(name):
    """(channel, clean_name) from a raw customer string with an optional prefix."""
    if not name or not isinstance(name, str):
        return None, (name.strip() if isinstance(name, str) else None)
    s = name.strip()
    m = CHANNEL_PREFIX.match(s)
    if m:
        ch = m.group(1).upper()
        ch = {"WSP": "WA", "WHATSAPP": "WA"}.get(ch, ch)
        clean = s[m.end():].strip()
        return ch, clean
    if s.startswith("#"):
        return "other", re.sub(r"^#\s*\d*\s*", "", s).strip()
    return "other", s

def norm_header(h):
    return re.sub(r"\s+", " ", str(h).strip()).upper() if h is not None else ""

def colmap(header):
    """normalized-header -> index, for the real (non-phantom) columns."""
    m = {}
    for i, h in enumerate(header):
        nh = norm_header(h)
        if not nh or (nh.startswith("COLUMNA") and nh[7:].isdigit()):
            continue
        m.setdefault(nh, i)
    return m

def pick(row, cm, *keys):
    for k in keys:
        i = cm.get(k)
        if i is not None and i < len(row):
            return row[i]
    return None

def rawjson(header, row, cm):
    d = {}
    for nh, i in cm.items():
        if i < len(row):
            v = row[i]
            if isinstance(v, (datetime, date)):
                v = v.isoformat()
            d[nh] = v
    return json.dumps(d, ensure_ascii=False, default=str)

# ---------- openpyxl ----------
import openpyxl

def sheet_rows(path, sheet=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = [sheet] if sheet else wb.sheetnames
    for name in names:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        it = ws.iter_rows(values_only=True)
        try:
            header = next(it)
        except StopIteration:
            continue
        cm = colmap(header)
        yield name, header, cm, list(it)
    wb.close()

# ---------- ingest: Provincias (invoice grain) ----------

def ingest_provincias(cur, downloads, report):
    files = sorted(glob.glob(os.path.join(downloads, "Provincias*.xlsx")))
    for path in files:
        base = os.path.basename(path)
        for name, header, cm, rows in sheet_rows(path, "Hoja1"):
            kept = dropped = 0
            for row in rows:
                iso = as_date(pick(row, cm, "FECHA"))
                if not iso:            # summary rows ("Limite mensual…") have no date
                    dropped += 1
                    continue
                raw_name = pick(row, cm, "NOMBRE")
                channel, clean = split_channel(raw_name)
                inv = pick(row, cm, "N° FACTURA", "Nº FACTURA", "N FACTURA")
                inv = str(inv).strip() if inv not in (None, "") else None
                iibb = num(pick(row, cm, "RETENCION IIBB"))
                sirtac = num(pick(row, cm, "IMP SIRTAC"))
                tax_total = (iibb or 0) + (sirtac or 0) if (iibb is not None or sirtac is not None) else None
                cur.execute(
                    """INSERT OR IGNORE INTO transactions
                    (source,channel,grain,date,year,customer_raw,customer_name,
                     gross,mp_fee,iibb,sirtac,tax_total,shipping,net,locality,invoice_number,
                     flags,raw_json,ingested_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    ("provincias", channel or "ML", "invoice", iso, year_of(iso),
                     raw_name, clean,
                     num(pick(row, cm, "PAGO")), num(pick(row, cm, "RECARGO MP")),
                     iibb, sirtac, tax_total,
                     num(pick(row, cm, "ENVIO")), num(pick(row, cm, "NETO")),
                     pick(row, cm, "LOCALIDAD"), inv,
                     date_flag(iso), rawjson(header, row, cm), NOW))
                kept += 1 if cur.rowcount else 0
                if not cur.rowcount:   # blocked by unique invoice index
                    report["dup_invoices"] += 1
            report["provincias"].append((base, kept, dropped))

# ---------- ingest: Máscaras (line_item grain) + expenses ----------

def ingest_mascaras(cur, downloads, report):
    path = os.path.join(downloads, "Máscaras Protectoras2.xlsx")
    if not os.path.exists(path):
        report["mascaras_missing"] = True
        return
    for name, header, cm, rows in sheet_rows(path):
        n = norm_header(name)
        if n in ("TABLA CONTROL", "PREDICCIONES"):   # derived pivots — skip
            continue
        if n == "GASTOS":
            kept = dropped = 0
            for row in rows:
                iso = as_date(pick(row, cm, "FECHA GASTO", "FECHA"))
                if not iso:
                    dropped += 1
                    continue
                cur.execute(
                    """INSERT INTO expenses
                    (source,date,year,item,qty,unit_cost,total_cost,payment_method,flags,raw_json,ingested_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                    ("mascaras", iso, year_of(iso),
                     pick(row, cm, "ITEM"), num(pick(row, cm, "CANTIDAD")),
                     num(pick(row, cm, "COSTO/UNIDAD")), num(pick(row, cm, "COSTO TOTAL")),
                     pick(row, cm, "FORMA DE PAGO"), date_flag(iso), rawjson(header, row, cm), NOW))
                kept += 1
            report["expenses"].append((name, kept, dropped))
            continue

        # sales sheets
        kept = dropped = 0
        for row in rows:
            iso = as_date(pick(row, cm, "FECHA"))
            if not iso:                     # trailing total rows have no date
                dropped += 1
                continue
            raw_name = pick(row, cm, "IDENTIFICACIÓN", "IDENTIFICACION")
            channel, clean = split_channel(raw_name)

            if "MASCARAS" in n:
                product = "Máscara"
                qty = (num(pick(row, cm, "CANT. ADULT.")) or 0) + (num(pick(row, cm, "CANT. NIÑOS")) or 0)
                unit = num(pick(row, cm, "PRECIO/UNIDAD"))
                gross = num(pick(row, cm, "PAGO")) or num(pick(row, cm, "EFT EN CAJA"))
            elif "BARBIJOS" in n:
                product = "Barbijo"
                qty = (num(pick(row, cm, "CANT. TB")) or 0) + (num(pick(row, cm, "CANT SUJ")) or 0)
                unit = num(pick(row, cm, "PRECIO/ UN TP", "PRECIO/UN TP"))
                gross = num(pick(row, cm, "EFT EN CAJA"))
            else:  # VENTA OTROS
                product = pick(row, cm, "ITEM")
                qty = num(pick(row, cm, "CANTIDAD"))
                unit = num(pick(row, cm, "PRECIO UNITARIO"))
                gross = num(pick(row, cm, "TOTAL"))

            cur.execute(
                """INSERT INTO transactions
                (source,channel,grain,date,year,customer_raw,customer_name,product,
                 qty,unit_price,gross,cost,profit,payment_method,notes,flags,raw_json,ingested_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                ("mascaras", channel, "line_item", iso, year_of(iso),
                 raw_name, clean, product,
                 qty or None, unit, gross,
                 num(pick(row, cm, "COSTO FABRICA")), num(pick(row, cm, "GANANCIAS")),
                 pick(row, cm, "PAGO REALIZADO"), pick(row, cm, "ENVIO RETIRO"),
                 date_flag(iso), rawjson(header, row, cm), NOW))
            kept += 1
        report["mascaras"].append((name, kept, dropped))

# ---------- report ----------

def money(x):
    return f"{x:,.0f}" if x is not None else "-"

def print_report(cur, report):
    p = print
    p("\n" + "=" * 70)
    p("INGEST SUMMARY")
    p("=" * 70)
    p("\nProvincias (invoice grain):")
    for base, kept, dropped in report["provincias"]:
        p(f"  {base:26} kept={kept:5}  dropped(summary rows)={dropped}")
    if report["dup_invoices"]:
        p(f"  duplicate invoice_number skipped: {report['dup_invoices']}")
    p("\nMáscaras sales (line_item grain):")
    for name, kept, dropped in report["mascaras"]:
        p(f"  {name:26} kept={kept:5}  dropped={dropped}")
    p("\nExpenses:")
    for name, kept, dropped in report["expenses"]:
        p(f"  {name:26} kept={kept:5}  dropped={dropped}")

    p("\n" + "-" * 70)
    p("TABLE TOTALS")
    for src, grain, cnt, dmin, dmax, g in cur.execute(
        """SELECT source, grain, COUNT(*), MIN(date), MAX(date), SUM(gross)
           FROM transactions GROUP BY source, grain ORDER BY source, grain"""):
        p(f"  transactions {src:11} {grain:10} rows={cnt:5}  {dmin}..{dmax}  gross={money(g)}")
    for cnt, dmin, dmax, t in cur.execute(
        "SELECT COUNT(*), MIN(date), MAX(date), SUM(total_cost) FROM expenses"):
        p(f"  expenses     {'mascaras':11} {'-':10} rows={cnt:5}  {dmin}..{dmax}  cost={money(t)}")

    p("\nChannel distribution (transactions):")
    for ch, cnt in cur.execute(
        "SELECT COALESCE(channel,'(null)'), COUNT(*) FROM transactions GROUP BY channel ORDER BY 2 DESC"):
        p(f"  {ch:10} {cnt}")

    p("\nDate-suspect rows flagged (kept, not dropped):")
    for r in cur.execute("SELECT year, COUNT(*) FROM transactions WHERE flags='date_suspect' GROUP BY year ORDER BY year"):
        p(f"  transactions year={r[0]}  rows={r[1]}")
    for r in cur.execute("SELECT year, COUNT(*) FROM expenses WHERE flags='date_suspect' GROUP BY year ORDER BY year"):
        p(f"  expenses     year={r[0]}  rows={r[1]}")

    p("\nInvoices per year (provincias):")
    for y, cnt, g in cur.execute(
        "SELECT year, COUNT(*), SUM(gross) FROM transactions WHERE source='provincias' GROUP BY year ORDER BY year"):
        p(f"  {y}  invoices={cnt:5}  gross={money(g)}")

    p("\nSample rows:")
    for r in cur.execute(
        "SELECT source,grain,date,channel,customer_name,product,gross,invoice_number "
        "FROM transactions WHERE source='provincias' LIMIT 2"):
        p("  ", r)
    for r in cur.execute(
        "SELECT source,grain,date,channel,customer_name,product,gross,invoice_number "
        "FROM transactions WHERE source='mascaras' AND product IS NOT NULL LIMIT 2"):
        p("  ", r)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--downloads", default=os.path.expanduser("~/Downloads"))
    ap.add_argument("--out", default=os.path.join(HERE, "history.db"))
    args = ap.parse_args()

    if os.path.exists(args.out):
        os.remove(args.out)
    con = sqlite3.connect(args.out)
    cur = con.cursor()
    with open(os.path.join(HERE, "schema.sql"), encoding="utf-8") as f:
        cur.executescript(f.read())

    report = {"provincias": [], "mascaras": [], "expenses": [],
              "dup_invoices": 0, "mascaras_missing": False}
    ingest_provincias(cur, args.downloads, report)
    ingest_mascaras(cur, args.downloads, report)
    con.commit()

    # Re-apply manual date corrections so a rebuild keeps them (idempotent).
    fixes = os.path.join(HERE, "fixes_dates.sql")
    if os.path.exists(fixes):
        with open(fixes, encoding="utf-8") as f:
            cur.executescript(f.read())
        con.commit()

    print_report(cur, report)
    con.close()
    print(f"\nWrote {args.out}")

if __name__ == "__main__":
    main()
